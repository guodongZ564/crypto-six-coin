"""解析《加密因子打分规则表.xlsx》为结构化配置。

只解析结构化的部分：
- 「1-体系总览」的维度权重表、regime系数表、分档表——这几处是干净的
  单元格网格，可靠地按坐标/表头锚点解析。
- 「2-打分规则表」的因子元数据（维度/适用币种/权重/是否反指/原始规则文本）。

"归一化/阈值规则"这一列是自然语言（有的是"<20→+2"这种能正则解析的绝对阈值，
有的是"崩溃(衰退)→-1"这种纯定性描述，根本不是能自动编译成公式的东西），
所以这里只把原始规则文本原样保留在 FactorRule.threshold_text 里当文档，
真正"当日值→[-2,+2]方向分"的计算逻辑按因子手写在 core/normalize.py，用
FactorRule.name 作为查找 key。没在 normalize.py 里实现的因子，仍然计入
「规则表应有因子数」分母（覆盖率播报要用），只是拿不到有效因子分。
"""

import re
from dataclasses import dataclass, field

import openpyxl

RULE_TABLE_PATH = "config/加密因子打分规则表.xlsx"

ALL_ASSETS = ["BTC", "ETH", "SOL", "UNI", "LINK", "LTC"]
ALTCOINS = ["ETH", "SOL", "UNI", "LINK", "LTC"]
COIN_DIMENSIONS = ["技术面", "资金面", "链上/基本面", "衍生品情绪"]
REGIME_DIMENSIONS = ["宏观", "行业周期"]


@dataclass
class FactorRule:
    dimension: str
    name: str
    applicable_assets: list
    bullish_direction: str
    threshold_text: str
    score_range: str
    weight: float
    source: str
    automation: str
    update_freq: str
    note: str
    is_reverse: bool


@dataclass
class RegimeBand:
    low: float | None  # None = 无下界
    high: float | None  # None = 无上界
    label: str
    long_mult: float
    short_mult: float


@dataclass
class ActionBand:
    low: float | None
    high: float | None
    action: str
    meaning: str


@dataclass
class SystemConfig:
    coin_dimension_weights: dict  # {"技术面": 0.35, ...} 和=1.0
    regime_dimension_weights: dict  # {"宏观": 0.5, "行业周期": 0.5}
    regime_bands: list  # list[RegimeBand]，按 low 升序
    action_bands: list  # list[ActionBand]，按 low 升序


def _parse_pct_pairs(text: str) -> dict:
    pairs = re.findall(r"([一-龥/]+)\s*(\d+(?:\.\d+)?)%", text)
    return {name: float(pct) / 100 for name, pct in pairs}


def _parse_bound_token(token: str) -> float:
    token = token.strip().lstrip("+")
    return float(token)


def _split_range_cell(text: str) -> tuple:
    """'> +0.3' / '≥ +1.0' / '-0.3 ~ +0.3' / '< -0.3' / '≤ -1.0' -> (low, high)"""
    text = text.strip()
    if text.startswith((">", "≥")):
        return _parse_bound_token(text[1:]), None
    if text.startswith(("<", "≤")):
        return None, _parse_bound_token(text[1:])
    if "~" in text:
        lo, hi = text.split("~")
        return _parse_bound_token(lo), _parse_bound_token(hi)
    raise ValueError(f"无法解析的区间: {text!r}")


def _find_row_by_first_cell(ws, text: str) -> int:
    for row in ws.iter_rows():
        if row[1].value == text:  # B列
            return row[1].row
    raise ValueError(f"没找到锚点行: {text!r}")


def _load_system_config(ws) -> SystemConfig:
    weight_row_regime = ws["B12"].value
    weight_row_coin = ws["B13"].value
    regime_dimension_weights = _parse_pct_pairs(weight_row_regime)
    coin_dimension_weights = _parse_pct_pairs(weight_row_coin)

    header_row = _find_row_by_first_cell(ws, "regime_score 区间")
    regime_bands = []
    for r in range(header_row + 1, header_row + 4):
        low, high = _split_range_cell(ws.cell(row=r, column=2).value)
        label = ws.cell(row=r, column=3).value
        long_mult = float(re.search(r"[\d.]+", ws.cell(row=r, column=4).value).group())
        short_mult = float(re.search(r"[\d.]+", ws.cell(row=r, column=5).value).group())
        regime_bands.append(RegimeBand(low=low, high=high, label=label, long_mult=long_mult, short_mult=short_mult))

    action_header_row = _find_row_by_first_cell(ws, "合成分")
    action_bands = []
    for r in range(action_header_row + 1, action_header_row + 6):
        low, high = _split_range_cell(ws.cell(row=r, column=2).value)
        action = ws.cell(row=r, column=3).value
        meaning = ws.cell(row=r, column=4).value
        action_bands.append(ActionBand(low=low, high=high, action=action, meaning=meaning))

    return SystemConfig(
        coin_dimension_weights=coin_dimension_weights,
        regime_dimension_weights=regime_dimension_weights,
        regime_bands=regime_bands,
        action_bands=action_bands,
    )


def _parse_applicable_assets(text: str) -> list:
    text = (text or "").strip()
    if text == "全市场":
        return list(ALL_ASSETS)
    if text == "各币":
        return list(ALL_ASSETS)
    if "山寨" in text:
        return list(ALTCOINS)
    if "/" in text:
        parsed = [a for a in text.split("/") if a in ALL_ASSETS]
        if parsed:
            return parsed
    if text in ALL_ASSETS:
        return [text]
    return []


def _load_factor_rules(ws) -> list:
    rules = []
    header_seen = False
    for row in ws.iter_rows():
        values = [c.value for c in row]
        if not header_seen:
            if len(values) > 1 and values[1] == "维度":
                header_seen = True
            continue
        padded = (values + [None] * 12)[:12]
        _, dimension, name, applicable, direction, threshold, score_range, weight, source, automation, update_freq, note = padded
        if not dimension or not name:
            continue
        note = note or ""
        direction = direction or ""
        rules.append(FactorRule(
            dimension=dimension,
            name=name,
            applicable_assets=_parse_applicable_assets(applicable),
            bullish_direction=direction,
            threshold_text=threshold or "",
            score_range=score_range or "",
            weight=float(weight) if weight is not None else 0.0,
            source=source or "",
            automation=automation or "",
            update_freq=update_freq or "",
            note=note,
            is_reverse=("反指" in note) or ("反指" in direction),
        ))
    return rules


def load_rule_table(path: str = RULE_TABLE_PATH) -> tuple:
    """返回 (SystemConfig, list[FactorRule])。"""
    wb = openpyxl.load_workbook(path, data_only=True)
    system_config = _load_system_config(wb["1-体系总览"])
    factor_rules = _load_factor_rules(wb["2-打分规则表"])
    return system_config, factor_rules
